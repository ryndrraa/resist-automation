from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from resist_automation.excel.reader import read_workbook_mapping
from resist_automation.excel.writer import export_project_to_excel
from resist_automation.exceptions import ExportError
from resist_automation.models import AxisResult
from resist_automation.rsx.mapper import scenario_from_parsed
from resist_automation.rsx.parser import parse_rsx
from resist_automation.utils.file_utils import sha256_file


def _complete_scenario(sample_rsx: Path, template_workbook: Path):
    scenario = scenario_from_parsed(
        parse_rsx(sample_rsx),
        read_workbook_mapping(template_workbook),
    )
    scenario.result_x = AxisResult(
        drift_ultimate=38,
        braced_ultimate=70,
        drift_service=46,
        braced_service=14,
    )
    scenario.result_y = AxisResult(
        drift_ultimate=37,
        braced_ultimate=68,
        drift_service=45,
        braced_service=14,
    )
    scenario.refresh_status()
    return scenario


def test_writer_places_values_and_preserves_template(
    sample_rsx: Path,
    template_workbook: Path,
    tmp_path: Path,
) -> None:
    scenario = _complete_scenario(sample_rsx, template_workbook)
    template_hash = sha256_file(template_workbook)
    original = load_workbook(template_workbook)
    original_style = original["PGA 0,5"]["G11"].style_id
    original.close()
    output = tmp_path / "hasil.xlsx"

    report = export_project_to_excel(template_workbook, output, [scenario])

    assert report.exported_scenarios == 1
    assert len(report.changes) == 10
    assert sha256_file(template_workbook) == template_hash
    workbook = load_workbook(output, data_only=False)
    worksheet = workbook["PGA 0,5"]
    expected = {
        "G11": 10,
        "L11": 10,
        "O11": 38,
        "P11": 70,
        "S11": 46,
        "T11": 14,
        "W11": 37,
        "X11": 68,
        "AA11": 45,
        "AB11": 14,
    }
    assert {cell: worksheet[cell].value for cell in expected} == expected
    assert worksheet["F11"].value == "=CONCATENATE(C11,D11,E11)"
    assert worksheet["G11"].style_id == original_style
    assert workbook["Sheet1"]["D4"].value == "=C4/((3^(1/2)*2))"
    workbook.close()


def test_writer_rejects_incomplete_results(
    sample_rsx: Path,
    template_workbook: Path,
    tmp_path: Path,
) -> None:
    scenario = scenario_from_parsed(
        parse_rsx(sample_rsx),
        read_workbook_mapping(template_workbook),
    )
    with pytest.raises(ExportError, match="belum lengkap"):
        export_project_to_excel(template_workbook, tmp_path / "hasil.xlsx", [scenario])
