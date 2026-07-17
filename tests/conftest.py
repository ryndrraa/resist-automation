from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill


@pytest.fixture
def sample_rsx() -> Path:
    return Path(__file__).parent / "fixtures" / "0.5 BFCTST-BFCTST.rsx"


@pytest.fixture
def template_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "Sheet1"
    sheet1["B1"] = 3.06
    sheet1["A4"] = "Bangunan Uji"
    sheet1["B4"] = 6
    sheet1["C4"] = 5
    sheet1["D4"] = "=C4/((3^(1/2)*2))"

    combinations = [
        (("BF", "CO", "ST"), ("BF", "CO", "ST")),
        (("BF", "CT", "ST"), ("BF", "CT", "ST")),
        (("BF", "EC", "ST"), ("BF", "EC", "ST")),
        (("BF", "CO", "ST"), ("BF", "CT", "ST")),
        (("BF", "CO", "ST"), ("BF", "EC", "ST")),
        (("BF", "CT", "ST"), ("BF", "EC", "ST")),
        (("BF", "EC", "ST"), ("BF", "CO", "ST")),
        (("BF", "CT", "ST"), ("BF", "CO", "ST")),
        (("BF", "EC", "ST"), ("BF", "CT", "ST")),
    ]
    for pga, title in ((0.4, "PGA 0,4"), (0.5, "PGA 0,5")):
        worksheet = workbook.create_sheet(title)
        for row, (x_codes, y_codes) in enumerate(combinations, start=10):
            worksheet[f"A{row}"] = row - 9
            worksheet[f"B{row}"] = f'="{pga:.1f} "&F{row}&"-"&K{row}'
            for column, value in zip(("C", "D", "E"), x_codes, strict=True):
                worksheet[f"{column}{row}"] = value
            worksheet[f"F{row}"] = f"=CONCATENATE(C{row},D{row},E{row})"
            for column, value in zip(("H", "I", "J"), y_codes, strict=True):
                worksheet[f"{column}{row}"] = value
            worksheet[f"K{row}"] = f"=CONCATENATE(H{row},I{row},J{row})"
            worksheet[f"G{row}"].fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    workbook.save(path)
    workbook.close()
    return path
