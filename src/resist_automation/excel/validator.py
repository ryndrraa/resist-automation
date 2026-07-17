from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from ..constants import REQUIRED_SHEETS, SCENARIO_ROWS
from ..exceptions import WorkbookError
from ..models import ValidationIssue, ValidationReport
from .reader import read_workbook_mapping


def validate_workbook(path: Path) -> ValidationReport:
    issues: list[ValidationIssue] = []
    try:
        read_workbook_mapping(path)
    except WorkbookError as exc:
        return ValidationReport(
            valid=False,
            issues=[ValidationIssue(severity="ERROR", code="WORKBOOK_INVALID", message=str(exc))],
        )

    workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        for sheet in REQUIRED_SHEETS:
            if sheet not in workbook.sheetnames:
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        code="MISSING_SHEET",
                        message=f"Sheet {sheet} tidak ditemukan.",
                    )
                )
        formula_cells = [("Sheet1", "D4")]
        for sheet in ("PGA 0,4", "PGA 0,5"):
            for row in SCENARIO_ROWS:
                formula_cells.extend((sheet, f"{column}{row}") for column in ("B", "F", "K"))
        for sheet, coordinate in formula_cells:
            value = workbook[sheet][coordinate].value
            if not (isinstance(value, str) and value.startswith("=")):
                issues.append(
                    ValidationIssue(
                        severity="ERROR",
                        code="MISSING_FORMULA",
                        message=f"Formula wajib {sheet}!{coordinate} tidak tersedia.",
                    )
                )
    finally:
        workbook.close()
    return ValidationReport(
        valid=not any(issue.severity == "ERROR" for issue in issues),
        issues=issues,
    )
