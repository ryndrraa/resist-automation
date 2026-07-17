from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..constants import PGA_SHEETS, REQUIRED_SHEETS, SCENARIO_ROWS
from ..exceptions import InvalidWorkbookError, MappingError
from ..models import MappingEntry, WorkbookMapping


def _join_codes(values: list[object], *, sheet: str, row: int, axis: str) -> str:
    parts = [str(value).strip() if value is not None else "" for value in values]
    if any(not part for part in parts):
        raise MappingError(
            f"Kode struktur sumbu {axis} pada {sheet} baris {row} tidak lengkap."
        )
    return "".join(parts).upper()


def read_workbook_mapping(path: Path) -> WorkbookMapping:
    path = Path(path).expanduser().resolve()
    if not path.exists():
        raise InvalidWorkbookError(f"Template workbook tidak ditemukan: {path}")
    if path.suffix.lower() != ".xlsx":
        raise InvalidWorkbookError("Template workbook harus berekstensi .xlsx.")
    try:
        workbook = load_workbook(path, data_only=False, read_only=False)
    except (InvalidFileException, OSError, ValueError) as exc:
        raise InvalidWorkbookError(f"Workbook {path.name} tidak dapat dibuka: {exc}") from exc

    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
    if missing:
        workbook.close()
        raise InvalidWorkbookError("Sheet wajib tidak ditemukan: " + ", ".join(missing))

    entries: list[MappingEntry] = []
    seen: set[tuple[str, str, str]] = set()
    try:
        for pga, sheet_name in PGA_SHEETS.items():
            worksheet = workbook[sheet_name]
            for row in SCENARIO_ROWS:
                structure_x = _join_codes(
                    [worksheet.cell(row, column).value for column in (3, 4, 5)],
                    sheet=sheet_name,
                    row=row,
                    axis="X",
                )
                structure_y = _join_codes(
                    [worksheet.cell(row, column).value for column in (8, 9, 10)],
                    sheet=sheet_name,
                    row=row,
                    axis="Y",
                )
                key = (sheet_name, structure_x, structure_y)
                if key in seen:
                    raise MappingError(
                        f"Kombinasi duplikat {structure_x}-{structure_y} pada sheet {sheet_name}."
                    )
                seen.add(key)
                entries.append(
                    MappingEntry(
                        sheet=sheet_name,
                        pga=pga,
                        row=row,
                        structure_x=structure_x,
                        structure_y=structure_y,
                    )
                )
    finally:
        workbook.close()
    return WorkbookMapping(template_path=path, entries=entries)
