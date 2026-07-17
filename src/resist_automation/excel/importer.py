from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from pydantic import ValidationError

from ..constants import RESULT_CELL_COLUMNS
from ..exceptions import InvalidWorkbookError
from ..models import AxisResult, Scenario


def _numeric_or_none(value: object, coordinate: str) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise InvalidWorkbookError(f"Nilai {coordinate} harus numerik atau kosong.")
    if value < 0:
        raise InvalidWorkbookError(f"Nilai {coordinate} tidak boleh negatif.")
    return float(value)


def import_results_from_workbook(path: Path, scenarios: list[Scenario]) -> int:
    path = Path(path).expanduser().resolve()
    if not path.exists() or path.suffix.lower() != ".xlsx":
        raise InvalidWorkbookError("Workbook hasil harus berupa file .xlsx yang tersedia.")
    try:
        workbook = load_workbook(path, data_only=True, read_only=False)
    except (OSError, ValueError) as exc:
        raise InvalidWorkbookError(f"Workbook hasil tidak dapat dibuka: {exc}") from exc

    imported = 0
    try:
        for scenario in scenarios:
            if not scenario.target_sheet or not scenario.target_row:
                continue
            if scenario.target_sheet not in workbook.sheetnames:
                continue
            worksheet = workbook[scenario.target_sheet]
            row = scenario.target_row
            values: dict[str, dict[str, float | None]] = {"x": {}, "y": {}}
            for axis in ("x", "y"):
                for field_name, column in RESULT_CELL_COLUMNS[axis].items():
                    coordinate = f"{column}{row}"
                    values[axis][field_name] = _numeric_or_none(
                        worksheet[coordinate].value,
                        f"{scenario.target_sheet}!{coordinate}",
                    )
            if not any(value is not None for axis in values.values() for value in axis.values()):
                continue
            try:
                scenario.result_x = AxisResult(**values["x"])
                scenario.result_y = AxisResult(**values["y"])
            except ValidationError as exc:
                raise InvalidWorkbookError(
                    f"Hasil untuk {scenario.scenario_id} tidak valid: {exc}"
                ) from exc
            scenario.refresh_status()
            imported += 1
    finally:
        workbook.close()
    return imported
