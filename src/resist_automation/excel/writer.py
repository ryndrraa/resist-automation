from __future__ import annotations

import logging
import os
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..constants import RESULT_CELL_COLUMNS, SCENARIO_ROWS
from ..exceptions import ExportError, OutputExistsError
from ..models import ExportCellChange, ExportReport, Scenario
from ..services.backup_service import create_backup
from .reader import read_workbook_mapping
from .validator import validate_workbook


LOGGER = logging.getLogger("resist_automation.excel.writer")


def _formula_snapshot(workbook: Any) -> dict[tuple[str, str], str]:
    formulas: dict[tuple[str, str], str] = {}
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas[(worksheet.title, cell.coordinate)] = cell.value
    return formulas


def _result_values(scenario: Scenario) -> list[float | None]:
    return [
        scenario.result_x.drift_ultimate,
        scenario.result_x.braced_ultimate,
        scenario.result_x.drift_service,
        scenario.result_x.braced_service,
        scenario.result_y.drift_ultimate,
        scenario.result_y.braced_ultimate,
        scenario.result_y.drift_service,
        scenario.result_y.braced_service,
    ]


def _validate_scenario(
    scenario: Scenario,
    *,
    allow_over_100: bool,
    occupied: set[tuple[str, int]],
) -> None:
    scenario.refresh_status()
    if scenario.status != "Lengkap":
        raise ExportError(
            f"Skenario {scenario.scenario_id} belum lengkap (status: {scenario.status})."
        )
    if scenario.target_sheet not in {"PGA 0,4", "PGA 0,5"}:
        raise ExportError(f"Target sheet skenario {scenario.scenario_id} tidak valid.")
    if scenario.target_row not in SCENARIO_ROWS:
        raise ExportError(f"Target baris skenario {scenario.scenario_id} harus 10 sampai 18.")
    target = (scenario.target_sheet, scenario.target_row)
    if target in occupied:
        raise ExportError(
            f"Lebih dari satu skenario menargetkan {scenario.target_sheet} baris {scenario.target_row}."
        )
    occupied.add(target)
    high_values = [value for value in _result_values(scenario) if value is not None and value > 100]
    if high_values and not allow_over_100:
        raise ExportError(
            f"Skenario {scenario.scenario_id} memiliki nilai di atas 100. "
            "Ekspor ulang dengan konfirmasi nilai tinggi."
        )


def _write_cell(
    worksheet: Any,
    coordinate: str,
    value: float | int,
    changes: list[ExportCellChange],
) -> None:
    old_value = worksheet[coordinate].value
    worksheet[coordinate] = value
    changes.append(
        ExportCellChange(
            sheet=worksheet.title,
            row=worksheet[coordinate].row,
            cell=coordinate,
            old_value=old_value,
            new_value=value,
        )
    )
    LOGGER.info(
        "action=write_excel sheet=%s row=%s cell=%s old=%r new=%r",
        worksheet.title,
        worksheet[coordinate].row,
        coordinate,
        old_value,
        value,
    )


def _verify_output(
    temp_path: Path,
    formulas: dict[tuple[str, str], str],
    changes: list[ExportCellChange],
) -> None:
    workbook = load_workbook(temp_path, data_only=False, read_only=False)
    try:
        for (sheet, coordinate), expected in formulas.items():
            if sheet not in workbook.sheetnames or workbook[sheet][coordinate].value != expected:
                raise ExportError(f"Formula berubah saat ekspor: {sheet}!{coordinate}.")
        for change in changes:
            actual = workbook[change.sheet][change.cell].value
            if actual != change.new_value:
                raise ExportError(
                    f"Verifikasi gagal pada {change.sheet}!{change.cell}: "
                    f"diharapkan {change.new_value!r}, ditemukan {actual!r}."
                )
    finally:
        workbook.close()


def export_project_to_excel(
    template_path: Path,
    output_path: Path,
    scenarios: list[Scenario],
    *,
    allow_overwrite: bool = False,
    allow_over_100: bool = False,
    backup_directory: Path | None = None,
) -> ExportReport:
    template_path = Path(template_path).expanduser().resolve()
    output_path = Path(output_path).expanduser().resolve()
    if output_path.suffix.lower() != ".xlsx":
        raise ExportError("File output harus berekstensi .xlsx.")
    if output_path == template_path and not allow_overwrite:
        raise ExportError("Output tidak boleh sama dengan template tanpa konfirmasi overwrite.")
    if output_path.exists() and not allow_overwrite:
        raise OutputExistsError(
            f"Output {output_path.name} sudah ada. Pilih nama baru atau konfirmasi overwrite."
        )
    if not scenarios:
        raise ExportError("Tidak ada skenario yang dapat diekspor.")

    workbook_report = validate_workbook(template_path)
    if not workbook_report.valid:
        messages = "; ".join(issue.message for issue in workbook_report.issues)
        raise ExportError(f"Template workbook tidak valid: {messages}")
    workbook_mapping = read_workbook_mapping(template_path)

    occupied: set[tuple[str, int]] = set()
    for scenario in scenarios:
        _validate_scenario(scenario, allow_over_100=allow_over_100, occupied=occupied)
        structure_x = scenario.structure_x.structure_id
        structure_y = scenario.structure_y.structure_id
        if structure_x and structure_y:
            expected = workbook_mapping.find(scenario.pga, structure_x, structure_y)
            if expected is None:
                raise ExportError(
                    f"Kombinasi {structure_x}-{structure_y} tidak ditemukan pada template."
                )
            if expected.sheet != scenario.target_sheet or expected.row != scenario.target_row:
                raise ExportError(
                    f"Mapping {scenario.scenario_id} tidak cocok dengan template: "
                    f"seharusnya {expected.sheet} baris {expected.row}."
                )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path: Path | None = None
    if allow_overwrite and output_path.exists():
        backup_path = create_backup(output_path, backup_directory)

    workbook = load_workbook(template_path, data_only=False, read_only=False)
    formulas = _formula_snapshot(workbook)
    changes: list[ExportCellChange] = []
    temp_handle = tempfile.NamedTemporaryFile(
        mode="wb",
        prefix=".resist-export-",
        suffix=".xlsx",
        dir=output_path.parent,
        delete=False,
    )
    temp_path = Path(temp_handle.name)
    temp_handle.close()
    try:
        for scenario in scenarios:
            assert scenario.target_sheet is not None and scenario.target_row is not None
            worksheet = workbook[scenario.target_sheet]
            row = scenario.target_row
            _write_cell(worksheet, f"G{row}", scenario.structure_x.depth_cm, changes)
            _write_cell(worksheet, f"L{row}", scenario.structure_y.depth_cm, changes)
            for axis, result in (("x", scenario.result_x), ("y", scenario.result_y)):
                for field_name, column in RESULT_CELL_COLUMNS[axis].items():
                    value = getattr(result, field_name)
                    assert value is not None
                    _write_cell(worksheet, f"{column}{row}", value, changes)
        workbook.save(temp_path)
        workbook.close()
        _verify_output(temp_path, formulas, changes)
        try:
            os.replace(temp_path, output_path)
        except PermissionError as exc:
            raise ExportError(
                f"Output {output_path.name} tidak dapat ditulis. Tutup file tersebut di Excel lalu coba lagi."
            ) from exc
    except Exception:
        workbook.close()
        temp_path.unlink(missing_ok=True)
        raise

    for scenario in scenarios:
        scenario.status = "Sudah diekspor"
    LOGGER.info(
        "action=export_complete source=%s output=%s scenarios=%s cells=%s",
        template_path,
        output_path,
        len(scenarios),
        len(changes),
    )
    return ExportReport(
        output_path=output_path,
        backup_path=backup_path,
        changes=changes,
        exported_scenarios=len(scenarios),
    )
